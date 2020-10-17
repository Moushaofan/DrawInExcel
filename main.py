import cv2 as cv
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def reverse(func):
    def inner(BGR):
        '''装饰器：反转BGR->RGB'''
        tmp = BGR[0]
        BGR[0] = BGR[2]
        BGR[2] = tmp
        return func(BGR)
    return inner
@reverse
def toHex(rgb):
    '''转换rgb ->16进制字符串'''
    tmp = rgb
    strs = ''
    for i in tmp:
        strs += str(hex(i))[-2:].replace('x','0').upper()  # 每次转换之后只取0x7b的后两位，拼接到strs中
    return strs

#"22017010012.jpg"
# 创建一个工作簿

print("开始填充")
def draw(filename,out):
    # 打开图片
    img = cv.imread(filename)
    if img is None:
        sys.exit("Could not read the image.")
    wb = Workbook()
    ws = wb.active  # 激活默认表格
    for i in range(img.shape[0]):
        for j in range(img.shape[1]):
            Color_str = toHex(img[i][j])
            print(i, j, img[i][j],Color_str)
            fill = PatternFill("solid", fgColor=Color_str)
            c = ws.cell(row=i + 1, column=j + 1)
            c.fill = fill
            print('填充完成 %d,%d'%(i,j))

    # 保存为一个xlsx格式的文件
    wb.save(out)
    # 关闭工作薄
    wb.close()
if __name__ == '__main__':
    draw('22017010012.jpg','draw3.xlsx')
